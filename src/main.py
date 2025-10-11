import json
import csv
import openpyxl
from typing import List, Dict
from search import process_bank_search

def load_json(filename: str) -> List[Dict]:
    with open(filename, encoding='utf-8') as f:
        return json.load(f)

def load_csv(filename: str) -> List[Dict]:
    with open(filename, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)

def load_xlsx(filename: str) -> List[Dict]:
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    keys = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in sheet.iter_rows(min_row=2):
        item = {k: c.value for k, c in zip(keys, row)}
        data.append(item)
    return data

def filter_by_status(data: List[Dict], status: str) -> List[Dict]:
    return [item for item in data if item.get('status') == status]

def sort_by_date(data: List[Dict], ascending: bool = True) -> List[Dict]:
    from datetime import datetime
    def parse_date(d):
        try:
            return datetime.strptime(d, "%d.%m.%Y")
        except Exception:
            return datetime.min
    return sorted(data, key=lambda x: parse_date(x.get('date', '')), reverse=not ascending)

def filter_by_currency(data: List[Dict], currency: str) -> List[Dict]:
    filtered = []
    for item in data:
        operation_sum = item.get('operationAmount') or item.get('amount') or item.get('sum')
        if not operation_sum:
            filtered.append(item)
            continue
        if isinstance(operation_sum, dict):
            cur = operation_sum.get('currency')
            if cur and cur.upper() == currency.upper():
                filtered.append(item)
        else:
            if isinstance(operation_sum, str) and currency.upper() in operation_sum.upper():
                filtered.append(item)
        return filtered
def user_input(prompt: str, values: List[str] = None) -> str:
    while True:
        answer = input(prompt).strip()
        if values is None or answer.lower() in [v.lower() for v in values]:
            return answer
        print(f'Ввод должен быть из: {", ".join(values)}')

def main():
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")

    file_type = user_input("Пользователь: ", ['1', '2', '3'])

    if file_type == '1':
        print("Программа: Для обработки выбран JSON-файл.")
        filename = input("Введите имя JSON-файла: ")
        data = load_json(filename)
    elif file_type == '2':
        print("Программа: Для обработки выбран CSV-файл.")
        filename = input("Введите имя CSV-файла: ")
        data = load_csv(filename)
    else:
        print("Программа: Для обработки выбран XLSX-файл.")
        filename = input("Введите имя XLSX-файла: ")
        data = load_xlsx(filename)

    valid_statuses = ['EXECUTED', 'CANCELED', 'PENDING']

    while True:
        status_input = input(
            f'Программа: Введите статус, по которому необходимо выполнить фильтрацию.\n'
            f'Доступные для фильтровки статусы: {", ".join(valid_statuses)}\nПользователь: ').strip()
        status_upper = status_input.upper()
        if status_upper in valid_statuses:
            print(f'Программа: Операции отфильтрованы по статусу "{status_upper}"')
            break
        print(f'Программа: Статус операции "{status_input}" недоступен.')

    filtered_data = [item for item in data if item.get('status') == status_upper]

    sort_answer = user_input('Программа: Отсортировать операции по дате? Да/Нет\nПользователь: ', ['Да', 'Нет'])
    if sort_answer.lower() == 'да':
        order_answer = user_input('Программа: Отсортировать по возрастанию или по убыванию?\nПользователь: ',
                                  ['по возрастанию', 'по убыванию'])
        ascending = order_answer == 'по возрастанию'
        filtered_data = sort_by_date(filtered_data, ascending)
    rub_answer = user_input('Программа: Выводить только рублевые транзакции? Да/Нет\nПользователь: ', ['Да', 'Нет'])
    if rub_answer.lower() == 'да':
        filtered_data = filter_by_currency(filtered_data, 'RUB')

    desc_answer = user_input(
        'Программа: Отфильтровать список транзакций по определенному слову в описании? Да/Нет\nПользователь: ',
        ['Да', 'Нет'])
    if desc_answer.lower() == 'да':
        search_str = input('Введите строку для поиска в описании:\nПользователь: ').strip()
        filtered_data = process_bank_search(filtered_data, search_str)

    print('Программа: Распечатываю итоговый список транзакций...')

    if not filtered_data:
        print('Программа: Не найдено ни одной транзакции, подходящей под ваши условия фильтрации')
        return

    print(f'Программа: Всего банковских операций в выборке: {len(filtered_data)}\n')

    for item in filtered_data:
        date = item.get('date', 'Неизвестна')
        description = item.get('description', 'Без описания')
        from_ = item.get('from', '')
        to = item.get('to', '')
        operation_sum = item.get('operationAmount') or item.get('amount') or item.get('sum')
        # Вывод суммы и валюты
        amount_str = ''
        if isinstance(operation_sum, dict):
            amount_str = f"{operation_sum.get('amount', '')} {operation_sum.get('currency', '')}".strip()
        elif isinstance(operation_sum, str):
            amount_str = operation_sum.strip()
        elif operation_sum is not None:
            amount_str = str(operation_sum).strip()

        print(f'{date} {description}')
        if from_ and to:
            print(f'{from_} -> {to}')
        elif from_:
            print(f'Откуда: {from_}')
        elif to:
            print(f'Куда: {to}')
        if amount_str:
            print(f'Сумма: {amount_str} \n')


if __name__ == '__main__':
    main()

